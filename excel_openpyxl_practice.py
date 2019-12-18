import openpyxl

print("********* Exccel Dataの読み込み********")
excel_in_path1 = './data/excel_in_header_2sheet.xlsx'
wb = openpyxl.load_workbook(excel_in_path1, read_only=False)

# シート名のリスト取得
sheet_names = wb.sheetnames
print(sheet_names) #['Member', 'History']

sheet1 = wb[sheet_names[0]]

# シート内の情報取得
# https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html
print(type(sheet1)) # <class 'openpyxl.worksheet.worksheet.Worksheet'>
print(sheet1['A1'].value) # セル番号を指定する方法
print(sheet1.cell(row=1, column=1).value)  # row/column指定は1行目,1列目が1になる（0ではあに）

print(sheet1.active_cell)

print("********* Exccel ブック操作********")
# https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.workbook.html

wb = openpyxl.Workbook()
# シートはdefaltで存在している
ws = wb.active
ws.title = "MySheet000"
print(wb.sheetnames)

# シートを３つ追加
wSheets = []
for index in range(1,4):
    print(index)
    wSheets.append(wb.create_sheet("MySheet{}".format(index)))

print(wb.sheetnames)

# シート名は変更できる
wSheets[0].title = "MySheet111"
print(wb.sheetnames)

# wbからforで回すことが可能
for sheet in wb:
    print(sheet.title)
    print(type(sheet)) #<class 'openpyxl.worksheet.worksheet.Worksheet'>

ws['A1'] = "From Python"
ws.cell(2,1, value="test1")
ws.cell(3,1).value = "test2"
ws.cell(6,1).value = "test5"


# 入力があるセルの最後までを表示
for cell_obj in list(ws.columns)[0]:
    print(cell_obj.value)

# シートをコピー(名前は 'MySheet000 Copy'と後ろにCopyがつく)
copy_ws = wb.copy_worksheet(ws)
print(wb.sheetnames)

# シートを一つ削除
wb.remove(wSheets[2])
print(wb.sheetnames)

print("********* cellの書式変更********")
# https://openpyxl.readthedocs.io/en/stable/styles.html
# https://openpyxl.readthedocs.io/en/stable/api/openpyxl.styles.fonts.html

from openpyxl.styles import Font, Color, colors
a1 = ws['A1']
ft = Font(color=colors.RED,
          name='Arial',
          size="24",
          underline='double')
a1.font = ft

# 罫線変更
from openpyxl.styles import Side, Border
side = Side(style='double', color='000000')
a1.border = Border(top=side, bottom=side, left=side, right=side)

# B2からE6までを罫線で引く(外側は実線、中は点線)
col_start = 2
row_start = 2
col_end = 5
row_end = 6

side_out = Side(style='thin', color='FF0000')
side_in =  Side(style='dashed', color='FF0000')

for col in range(col_start, col_end + 1):
    for row in range(row_start, row_end + 1):
        cell = ws.cell(row=row, column=col)

        # 順番的に最左列の上から下へ順場に実施していく
        if col == col_start:
            if row == row_start:
                # 左上
                cell.border = Border(top=side_out, bottom=side_in, left=side_out)
            elif row == row_end:
                # 左下
                cell.border = Border(bottom=side_out, left=side_out)
            else:
                # 残りの列
                cell.border = Border(bottom=side_in, left=side_out)

        elif col == col_end:
            if row == row_start:
                # 右上
                cell.border = Border(top=side_out, bottom=side_in, right=side_out)
            elif row == row_end:
                # 左下
                cell.border = Border(bottom=side_out, right=side_out)
            else:
                # 残りの列
                cell.border = Border(bottom=side_in, right=side_out)
        else:
            #その他の列
            if row == row_start:
                # 上の行
                cell.border = Border(top=side_out, bottom=side_in, left=side_in)
            elif row == row_end:
                # 左下
                cell.border = Border(bottom=side_out, left=side_in)
            else:
                # 残りの列
                cell.border = Border(bottom=side_in, left=side_in)

        # 実際は horizonやverticalを使った方が良さそう


# セル幅変更
ws.row_dimensions[1].height = 24
ws.column_dimensions['A'].width = 30

# 保存
wb.save('./data/excel_out_new_from_openpyxl.xlsx')
